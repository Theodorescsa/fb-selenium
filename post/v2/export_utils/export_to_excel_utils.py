import json
import os
from datetime import datetime
from typing import Dict, Any, List
from openpyxl import Workbook, load_workbook  # thêm import cho Excel

EXCEL_PATH = "data/thoibaode_db.xlsx"

# Các cột bắt buộc — bạn có thể chỉnh tùy theo cấu trúc dữ liệu của bạn
REQUIRED_COLUMNS = [
    "id",
    "title",
    "content",
    "created_time",
    "author",
    "url",
]

def _jsonable(val):
    """Convert list/dict sang chuỗi để Excel nuốt được."""
    if isinstance(val, (list, dict)):
        return json.dumps(val, ensure_ascii=False)
    return val

def _normalize_created_time(val):
    """created_time trong mẫu là epoch (int) → đổi sang yyyy-mm-dd hh:mm:ss."""
    if isinstance(val, int):
        return datetime.utcfromtimestamp(val).strftime("%Y-%m-%d %H:%M:%S")
    return val

def write_posts_to_excel(posts: List[Dict[str, Any]], excel_path: str = EXCEL_PATH):
    if not posts:
        return

    # 1) Nếu đã có file → mở, không thì tạo mới
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
        current_headers = [c.value for c in ws[1] if c.value]
    else:
        wb = Workbook()
        ws = wb.active
        current_headers = []

    # 2) Tập header cuối cùng
    final_headers = list(current_headers) if current_headers else list(REQUIRED_COLUMNS)

    for col in REQUIRED_COLUMNS:
        if col not in final_headers:
            final_headers.append(col)

    extra_keys = set()
    for p in posts:
        extra_keys.update(p.keys())
    for k in sorted(extra_keys):
        if k not in final_headers:
            final_headers.append(k)

    # 3) Nếu header thay đổi → ghi lại hàng 1
    if final_headers != current_headers:
        for idx, h in enumerate(final_headers, start=1):
            ws.cell(row=1, column=idx, value=h)

    # 4) Ghi từng post
    start_row = ws.max_row + 1
    for p in posts:
        for col_idx, col_name in enumerate(final_headers, start=1):
            val = p.get(col_name)
            if col_name == "created_time":
                val = _normalize_created_time(val)
            val = _jsonable(val)
            ws.cell(row=start_row, column=col_idx, value=val)
        start_row += 1

    wb.save(excel_path)

import time
import json
import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from seleniumwire import webdriver
from selenium.webdriver.chrome.options import Options

# ====== import từ code của ông ======
from get_comment_fb_utils import (
    open_reel_comments_if_present,
    set_sort_to_all_comments_unified,
)
from get_comment_fb_automation import (
    bootstrap_auth,
    hook_graphql,
    install_early_hook,
)
from main import crawl_comments  # nếu file ông tên khác thì đổi lại
from configs import *

# ====== cấu hình ======
PROXY_URL = ""
INPUT_EXCEL = r"E:\NCS\fb-selenium\thoibao-de-last.xlsx"          # file excel nguồn (có cột link)
OUTPUT_EXCEL = r"E:\NCS\fb-selenium\thoibaode-comments.xlsx"      # file excel đích
ERROR_EXCEL = r"E:\NCS\fb-selenium\crawl_errors.xlsx"             # file ghi những post crawl fail
CHECKPOINT_PATH = r"E:\NCS\fb-selenium\crawl_checkpoint.json"     # base checkpoint
SHEET_NAME = "Sheet1"  # 👈 👈 👈 THÊM Ở ĐÂY: chỉ định sheet muốn crawl
MAX_RETRIES = 2                                                   # retry tối đa

# cột mong muốn cho file output
OUTPUT_COLUMNS = [
    "id",
    "type",
    "postlink",
    "commentlink",
    "author_id",
    "author",
    "author_link",
    "avatar",
    "created_time",
    "content",
    "image_url",
    "like",
    "comment",
    "haha",
    "wow",
    "sad",
    "love",
    "angry",
    "care",
    "video",
    "source_id",
    "is_share",
    "link_share",
    "type_share",
]


def start_driver_with_proxy(proxy_url: str, headless: bool = False) -> webdriver.Chrome:
    chrome_opts = Options()
    if headless:
        chrome_opts.add_argument("--headless=new")
        chrome_opts.add_argument("--disable-gpu")
    chrome_opts.add_argument("--no-sandbox")
    chrome_opts.add_argument("--disable-dev-shm-usage")
    chrome_opts.add_argument("--window-size=1920,1080")
    chrome_opts.add_argument("--disable-extensions")
    chrome_opts.add_argument("--disable-background-networking")
    chrome_opts.add_argument("--disable-popup-blocking")
    chrome_opts.add_argument("--no-first-run")
    chrome_opts.add_argument("--no-default-browser-check")
    chrome_opts.add_argument("--disable-background-timer-throttling")
    chrome_opts.add_argument("--disable-backgrounding-occluded-windows")
    chrome_opts.add_argument("--disable-renderer-backgrounding")

    sw_options = None
    if proxy_url:
        sw_options = {
            "proxy": {
                "http": proxy_url,
                "https": proxy_url,
                "no_proxy": "localhost,127.0.0.1",
            },
        }

    driver = webdriver.Chrome(options=chrome_opts, seleniumwire_options=sw_options)
    driver.scopes = [r".*"]
    return driver


def ensure_excel_with_header(path: str, columns: list[str]):
    """Tạo file excel nếu chưa có và ghi header."""
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append(columns)
        wb.save(path)


def append_row_to_excel(path: str, columns: list[str], row_dict: dict):
    wb = load_workbook(path)
    ws = wb.active
    row = []
    for col in columns:
        v = row_dict.get(col)
        if isinstance(v, dict):
            v = v.get("uri") or v.get("url") or json.dumps(v, ensure_ascii=False)
        elif isinstance(v, list):
            v = ",".join(str(x) for x in v)
        row.append(v)
    ws.append(row)
    wb.save(path)


def ensure_error_excel(path: str):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append(["link", "error"])
        wb.save(path)


def append_error(path: str, link: str, error: str):
    wb = load_workbook(path)
    ws = wb.active
    ws.append([link, error])
    wb.save(path)


def make_checkpoint_path_for_sheet(base_path: str, sheet_name: str) -> str:
    """
    Tạo checkpoint riêng cho từng sheet.
    vd: crawl_checkpoint.json  -> crawl_checkpoint_Sheet1.json
    """
    base_dir = os.path.dirname(base_path)
    base_name = os.path.basename(base_path)
    name, ext = os.path.splitext(base_name)
    safe_sheet = sheet_name.replace(" ", "_")
    final_name = f"{name}_{safe_sheet}{ext}"
    return os.path.join(base_dir, final_name)


def load_checkpoint(path: str) -> int:
    """
    Trả về index bắt đầu crawl.
    - Nếu chưa có file → trả 0 (crawl từ đầu).
    - Nếu có → đọc "last_index" và +1 để crawl post tiếp theo.
    """
    if not os.path.exists(path):
        return 0
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        last_idx = int(data.get("last_index", -1))
        return last_idx + 1 if last_idx >= 0 else 0
    except Exception:
        return 0


def save_checkpoint(path: str, idx: int, total: int):
    data = {
        "last_index": idx,
        "total": total,
        "ts": time.time(),
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def prepare_fb_page(driver, url: str):
    """Mở post và chuẩn bị để replay GraphQL."""
    driver.get(url)
    time.sleep(2)
    hook_graphql(driver)
    time.sleep(0.5)
    if "reel" in url:
        open_reel_comments_if_present(driver)
    set_sort_to_all_comments_unified(driver)
    time.sleep(0.8)


def crawl_one_post(driver, url: str, max_pages=None):
    out_json = "comments_tmp.ndjson"
    ckpt = "checkpoint_tmp.json"

    # xoá file cũ
    if os.path.exists(out_json):
        os.remove(out_json)
    if os.path.exists(ckpt):
        os.remove(ckpt)

    prepare_fb_page(driver, url)

    _ = crawl_comments(
        driver,
        out_json=out_json,
        checkpoint_path=ckpt,
        max_pages=max_pages,
    )

    comments = []
    if os.path.exists(out_json):
        with open(out_json, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = json.loads(line)
                    comments.append(rec)
                except Exception:
                    pass

    return comments


def normalize_comment_row(c: dict, postlink: str) -> dict:
    _type = c.get("type") or ("Reply" if c.get("is_reply") else "Comment")
    comment_link = c.get("link")

    return {
        "id": c.get("id") or c.get("reply_id"),
        "type": _type,
        "postlink": postlink,
        "commentlink": comment_link,
        "author_id": c.get("author_id"),
        "author": c.get("author"),
        "author_link": c.get("author_link"),
        "avatar": c.get("avatar"),
        "created_time": c.get("created_time"),
        "content": c.get("content") or c.get("text"),
        "image_url": ",".join(c.get("image_url") or []) if isinstance(c.get("image_url"), list) else c.get("image_url"),
        "like": c.get("like", 0),
        "comment": c.get("comment", 0),
        "haha": c.get("haha", 0),
        "wow": c.get("wow", 0),
        "sad": c.get("sad", 0),
        "love": c.get("love", 0),
        "angry": c.get("angry", 0),
        "care": c.get("care", 0),
        "video": ",".join(c.get("video") or []) if isinstance(c.get("video"), list) else c.get("video"),
        "source_id": c.get("source_id"),
        "is_share": c.get("is_share"),
        "link_share": c.get("link_share"),
        "type_share": c.get("type_share"),
    }


def crawl_from_excel_stream(
    input_path: str,
    sheet_name: str,
    output_path: str,
    error_path: str,
    checkpoint_path: str,
    driver,
    max_retries: int = 2,
):
    ensure_excel_with_header(output_path, OUTPUT_COLUMNS)
    ensure_error_excel(error_path)

    # 👇 đọc đúng sheet
    df = pd.read_excel(input_path, sheet_name=sheet_name)
    total = len(df)

    # 👇 tạo checkpoint riêng cho sheet này
    sheet_checkpoint = make_checkpoint_path_for_sheet(checkpoint_path, sheet_name)
    start_idx = load_checkpoint(sheet_checkpoint)

    if start_idx > 0:
        print(f"🔁 Resume từ dòng {start_idx} / {total} (sheet={sheet_name})")
    else:
        print(f"🆕 Chạy mới từ đầu (sheet={sheet_name})")

    for idx in range(start_idx, total):
        row = df.iloc[idx]
        postlink = str(row.get("link") or "").strip()
        if not postlink:
            save_checkpoint(sheet_checkpoint, idx, total)
            continue

        print(f"=== [{idx+1}/{total}] ({sheet_name}) Crawl: {postlink}")

        success = False
        last_error = None

        for attempt in range(max_retries + 1):
            try:
                comments = crawl_one_post(driver, postlink, max_pages=None)
                for c in comments:
                    norm = normalize_comment_row(c, postlink)
                    append_row_to_excel(output_path, OUTPUT_COLUMNS, norm)
                success = True
                break
            except Exception as e:
                last_error = str(e)
                print(f"[WARN] crawl fail {postlink} (attempt {attempt+1}/{max_retries+1}): {e}")
                time.sleep(1)

        # ✅ Dù ok hay fail vẫn lưu checkpoint
        if not success:
            append_error(error_path, postlink, last_error or "unknown error")
            print(f"[SKIP] bỏ qua bài: {postlink}")

        save_checkpoint(sheet_checkpoint, idx, total)

    print(f"✅ DONE sheet {sheet_name} → xem file: {output_path}")


if __name__ == "__main__":
    d = start_driver_with_proxy(PROXY_URL, headless=False)
    d.set_script_timeout(40)
    try:
        d.execute_cdp_cmd("Network.enable", {})
        d.execute_cdp_cmd("Network.setCacheDisabled", {"cacheDisabled": True})
    except Exception:
        pass

    bootstrap_auth(d)
    install_early_hook(d)

    crawl_from_excel_stream(
        INPUT_EXCEL,
        SHEET_NAME,          # 👈 truyền sheet muốn crawl
        OUTPUT_EXCEL,
        ERROR_EXCEL,
        CHECKPOINT_PATH,
        driver=d,
        max_retries=MAX_RETRIES,
    )
    # d.quit()

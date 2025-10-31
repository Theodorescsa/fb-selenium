import os
import time
import json
import hashlib
from pathlib import Path
from typing import Optional, List, Dict, Any

import pandas as pd
from openpyxl import Workbook, load_workbook

# ==== import từ code GỐC của bạn (KHÔNG sửa mấy file này) ====
# Lưu ý: module "main" phải có crawl_comments & crawl_replies_for_parent_expansion
#        module "get_comment_fb_automation" phải có start_driver, hook, ...
from get_comment_fb_utils import open_reel_comments_if_present, set_sort_to_all_comments_unified
from get_comment_fb_automation import (
    start_driver,
    install_early_hook,
    hook_graphql,
)
import main as _core  # <- file logic gốc bạn vừa gửi (chứa crawl_comments & crawl_replies_for_parent_expansion)

# =========================
# CONFIG (điều chỉnh theo máy bạn)
# =========================
HERE = Path(__file__).resolve().parent
DATABASE_PATH = Path(__file__).resolve().parent.parent.parent / "database"
INPUT_EXCEL = DATABASE_PATH / "post" / "page" / "thoibaode" / "thoibao-de-last-split-sheet1.xlsx"
PROXY_URL = ""
SHEET_NAME = "Sheet_1"
OUTPUT_EXCEL = DATABASE_PATH / "comment" / "page" / "thoibaode" / "sheet1" / "thoibaode-comments-sheet1.xlsx"     
ERROR_EXCEL = DATABASE_PATH / "comment" / "page" / "thoibaode" / "sheet1" / "crawl_errors-sheet1.xlsx"           
STATUS_COL_NAME = "status"                                              
STATUS_STORE_PATH = DATABASE_PATH / "comment" / "page" / "thoibaode" / "sheet1" / "status_store_sheet1.json"

# Temp directory cho NDJSON & checkpoint (per-post)
TMP_DIR = DATABASE_PATH / "comment" / "page" / "thoibaode" / "sheet1" / "tmp_comments_sheet1"
os.makedirs(TMP_DIR, exist_ok=True)

# Cache JSON chống crawl replies lặp theo post
DEDUP_CACHE_PATH = DATABASE_PATH / "comment" / "page" / "thoibaode" / "sheet1" / "reply_dedup_cache_sheet1.json"

# Chrome profile (từ configs.py của bạn)
try:
    from configs import CHROME_PATH, USER_DATA_DIR, PROFILE_NAME, REMOTE_PORT
except Exception:
    # fallback: điền tay nếu cần
    CHROME_PATH = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    USER_DATA_DIR = r"E:\NCS\ChromeProfiles"
    PROFILE_NAME = "Default"
    REMOTE_PORT = 9222

# Cột output cho file kết quả
OUTPUT_COLUMNS = [
    "id", "type", "postlink", "commentlink", "author_id", "author", "author_link",
    "avatar", "created_time", "content", "image_url", "like", "comment", "haha",
    "wow", "sad", "love", "angry", "care", "video", "source_id", "is_share",
    "link_share", "type_share"
]

# =========================
# Status JSON helpers (thay vì ghi vào Excel)
# =========================
def _atomic_write_json(path: str, data: dict):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)

def load_status_store(path: str) -> dict[str, str]:
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                d = json.load(f) or {}
                # chuẩn hóa: chỉ giữ string
                return {str(k): str(v) for k, v in d.items()}
        except Exception:
            pass
    return {}

def save_status_store(path: str, store: dict[str, str]):
    try:
        _atomic_write_json(path, store)
    except Exception:
        # best-effort
        pass

def get_status(store: dict[str, str], postlink: str) -> str:
    return (store.get(postlink) or "").strip().lower()

def set_status(store: dict[str, str], postlink: str, status: str):
    store[postlink] = status
# =========================
# Excel helpers
# =========================
def ensure_excel_with_header(path: str, columns: list[str]):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append(columns)
        wb.save(path)

def ensure_error_excel(path: str):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append(["link", "error"])
        wb.save(path)

def append_row_to_excel(path: str, columns: list[str], row_dict: dict):
    wb = load_workbook(path)
    ws = wb.active
    row = []
    for col in columns:
        v = row_dict.get(col)
        if isinstance(v, dict):
            v = v.get("uri") or v.get("url") or json.dumps(v, ensure_ascii=False)
        elif isinstance(v, list):
            v = ",".join(map(str, v))
        row.append(v)
    ws.append(row)
    wb.save(path)

def append_error(path: str, link: str, error: str):
    wb = load_workbook(path)
    ws = wb.active
    ws.append([link, error])
    wb.save(path)

def ensure_status_col(input_path: str, sheet_name: str, status_col_name: str = "status") -> int:
    wb = load_workbook(input_path)
    ws = wb[sheet_name]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if status_col_name in header:
        col_idx = header.index(status_col_name) + 1
        wb.save(input_path)
        return col_idx
    ws.cell(row=1, column=len(header) + 1, value=status_col_name)
    wb.save(input_path)
    return len(header) + 1

def set_row_status(input_path: str, sheet_name: str, row_index_1based: int, status_value: str, status_col_idx: int):
    wb = load_workbook(input_path)
    ws = wb[sheet_name]
    ws.cell(row=row_index_1based, column=status_col_idx, value=status_value)
    wb.save(input_path)

def read_cell(input_path: str, sheet_name: str, row_index_1based: int, col_name: str):
    wb = load_workbook(input_path, read_only=True)
    ws = wb[sheet_name]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if col_name not in header:
        return None, None
    col_idx = header.index(col_name) + 1
    if row_index_1based > ws.max_row:
        return col_idx, None
    cell = ws.cell(row=row_index_1based, column=col_idx)
    return col_idx, (cell.value if cell else None)

# =========================
# Monkey-patch: chặn replies lặp (KHÔNG sửa file core)
# =========================
def _load_cache():
    if os.path.exists(DEDUP_CACHE_PATH):
        try:
            with open(DEDUP_CACHE_PATH, "r", encoding="utf-8") as f:
                return json.load(f) or {}
        except Exception:
            pass
    return {}

def _save_cache(data):
    try:
        with open(DEDUP_CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

_reply_cache = _load_cache()  # { post_url: { parent_id: ts } }
__orig_crawl_replies = _core.crawl_replies_for_parent_expansion

def _patched_crawl_replies_for_parent_expansion(
    driver,
    url,
    form,
    base_reply_vars,
    parent_id,
    parent_token,
    out_json,
    extract_fn,
    clean_fn,
    max_reply_pages=None
):
    # Dùng URL hiện tại nếu có (ổn định theo post), fallback url tham số
    post_key = str(driver.current_url or url)
    seen_for_post = _reply_cache.setdefault(post_key, {})

    if parent_id in seen_for_post:
        print(f"[PATCH] Skip replies for parent={parent_id[:12]}… (already processed for this post)")
        return

    __orig_crawl_replies(
        driver,
        url,
        form,
        base_reply_vars,
        parent_id,
        parent_token,
        out_json,
        extract_fn,
        clean_fn,
        max_reply_pages=max_reply_pages
    )

    seen_for_post[parent_id] = int(time.time())
    _save_cache(_reply_cache)

# Áp dụng patch ngay
_core.crawl_replies_for_parent_expansion = _patched_crawl_replies_for_parent_expansion
print("[PATCH] Installed replies de-dup (no changes in core files).")

# =========================
# Helpers
# =========================
def normalize_comment_row(c: dict, postlink: str) -> dict:
    _type = c.get("type") or ("Reply" if c.get("is_reply") else "Comment")
    comment_link = c.get("link")
    print("Trả lời:",c.get("content") or c.get("text") or c.get("message") or (c.get("body") if isinstance(c.get("body"), str) else None))
    return {
        "id": c.get("id") or c.get("raw_comment_id") or c.get("reply_id"),
        "type": _type,
        "postlink": postlink,
        "commentlink": comment_link,
        "author_id": c.get("author_id"),
        "author": c.get("author"),
        "author_link": c.get("author_link"),
        "avatar": c.get("avatar"),
        "created_time": c.get("created_time"),
        "content": c.get("content") or c.get("text") or c.get("message") or (c.get("body") if isinstance(c.get("body"), str) else None),
        "image_url": ",".join(c.get("image_url") or []) if isinstance(c.get("image_url"), list) else c.get("image_url"),
        "like": c.get("like", 0),
        "comment": c.get("comment", 0),
        "haha": c.get("haha", 0),
        "wow": c.get("wow", 0),
        "sad": c.get("sad", 0),
        "love": c.get("love", 0),
        "angry": c.get("angry", 0),
        "care": c.get("care", 0),
        "video": ",".join(c.get("video") or []) if isinstance(c.get("video"), list) else c.get("video"),
        "source_id": c.get("source_id"),
        "is_share": c.get("is_share"),
        "link_share": c.get("link_share"),
        "type_share": c.get("type_share"),
    }

def build_post_temp_paths(postlink: str) -> tuple[str, str]:
    """Sinh đường dẫn out_json & checkpoint riêng theo postlink (hash)"""
    h = hashlib.md5(postlink.encode("utf-8")).hexdigest()[:16]
    out_json = os.path.join(TMP_DIR, f"comments_{h}.ndjson")
    ckpt = os.path.join(TMP_DIR, f"checkpoint_{h}.json")
    return out_json, ckpt

def load_ndjson(path: str) -> List[Dict[str, Any]]:
    out = []
    if not os.path.exists(path):
        return out
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                out.append(json.loads(line))
            except Exception:
                pass
    return out

def read_existing_pairs_in_output(path: str) -> set[tuple[str, str]]:
    """Đọc nhanh cặp (postlink, id) đã có trong OUTPUT_EXCEL để tránh duplicate."""
    pairs = set()
    if not os.path.exists(path):
        return pairs
    try:
        df = pd.read_excel(path, usecols=["postlink", "id"])
        for _, r in df.iterrows():
            pl = str(r.get("postlink") or "").strip()
            cid = str(r.get("id") or "").strip()
            if pl and cid:
                pairs.add((pl, cid))
    except Exception:
        # fallback openpyxl nếu muốn, nhưng đủ dùng với pandas
        pass
    return pairs

# =========================
# Main crawl loop (Excel)
# =========================
# đặt gần đầu file (trước crawl_one_post) — exception + helper siêu nhẹ
class BlockedInCountryError(Exception):
    pass

def _is_blocked_in_country(driver, timeout_sec: float = 2.5) -> str | None:
    """
    Trả về lý do nếu bài bị chặn theo quốc gia; None nếu không bị chặn.
    Chỉ nhìn page_source 2–3s, không hook/scroll cho nhanh.
    """
    import time
    needles = [
        "Bài viết này không hiển thị tại Việt Nam",
        "Do chúng tôi đáp ứng yêu cầu từ Vietnam Ministry of Culture, Sports and Tourism",
        "This content isn't available in your country",
        "This post isn't available in your country",
        "not available in your country",
    ]
    t0 = time.time()
    while time.time() - t0 < timeout_sec:
        src = (driver.page_source or "").lower()
        for n in needles:
            if n.lower() in src:
                return n
        time.sleep(0.2)
    return None


def crawl_one_post(driver, postlink: str, max_pages=None) -> List[Dict[str, Any]]:
    out_json, ckpt = build_post_temp_paths(postlink)
    # dọn tệp cũ để tránh nhập nhằng (tuỳ bạn muốn giữ resume thì comment 2 dòng dưới)
    if os.path.exists(out_json):
        os.remove(out_json)
    if os.path.exists(ckpt):
        os.remove(ckpt)

    # 👉 mở link và precheck block NGAY TẠI ĐÂY
    driver.get(postlink)
    time.sleep(0.8)  # đủ để banner render
    reason = _is_blocked_in_country(driver, timeout_sec=2.5)
    if reason:
        # ném lỗi nhẹ để vòng ngoài đánh fail & skip
        raise BlockedInCountryError(f"blocked_in_vietnam: {reason}")

    # không bị chặn thì mới hook graphql, set sort, ...
    time.sleep(0.4)
    hook_graphql(driver)
    time.sleep(0.4)
    if "reel" in postlink:
        try:
            open_reel_comments_if_present(driver)
        except Exception:
            pass
    try:
        set_sort_to_all_comments_unified(driver)
    except Exception:
        pass
    time.sleep(0.6)

    # gọi core crawl (vẫn dùng tệp tạm)
    texts = _core.crawl_comments(
        driver,
        out_json=out_json,
        checkpoint_path=ckpt,
        max_pages=max_pages
    )

    # ghép: đọc lại NDJSON (ổn định hơn do core “append từng dòng”)
    return load_ndjson(out_json)

def crawl_from_excel_stream(
    input_path: str,
    sheet_name: str,
    output_path: str,
    error_path: str,
    driver,
    max_retries: int = 0,
):
    ensure_excel_with_header(output_path, OUTPUT_COLUMNS)
    ensure_error_excel(error_path)

    # Load status store JSON (link -> status)
    status_store = load_status_store(STATUS_STORE_PATH)

    # đọc sheet (pandas)
    df = pd.read_excel(input_path, sheet_name=sheet_name)
    total = len(df)

    # đọc cặp đã có trong OUTPUT_EXCEL để tránh duplicate
    existing_pairs = read_existing_pairs_in_output(output_path)

    print(f"▶️ Bắt đầu crawl (sheet={sheet_name}), tổng {total} dòng")
    for i in range(total):
        postlink = str(df.iloc[i].get("link") or "").strip()
        if not postlink:
            # không có link -> fail (nhưng fail này ko có link để log)
            continue

        # kiểm tra status hiện tại trong JSON
        current_status = get_status(status_store, postlink)

        if current_status == "done":
            print(f"⏩ [{i+1}/{total}] SKIP (done): {postlink}")
            continue

        print(f"=== [{i+1}/{total}] Crawl: {postlink}")
        success = False
        last_error = None

        for attempt in range(max_retries + 1):
            try:
                records = crawl_one_post(driver, postlink, max_pages=None)
                # normalize + de-dup theo (postlink, id)
                new_cnt = 0
                for c in records:
                    norm = normalize_comment_row(c, postlink)
                    cid = str(norm.get("id") or "").strip()
                    if not cid:
                        continue
                    key = (postlink, cid)
                    if key in existing_pairs:
                        continue
                    append_row_to_excel(output_path, OUTPUT_COLUMNS, norm)
                    existing_pairs.add(key)
                    new_cnt += 1
                print(f"[WRITE] {new_cnt} new rows → {output_path}")
                success = True
                break
            except BlockedInCountryError as e:
                last_error = str(e)
                print(f"[BLOCKED] {postlink} → {last_error}")
                break
            except Exception as e:
                last_error = str(e)
                print(f"[WARN] crawl fail {postlink} (attempt {attempt+1}/{max_retries+1}): {e}")
                time.sleep(1)

        # ghi trạng thái vào JSON
        if success:
            set_status(status_store, postlink, "done")
        else:
            append_error(error_path, postlink, last_error or "unknown error")
            set_status(status_store, postlink, "fail")
            print(f"[SKIP] bỏ qua bài: {postlink}")

        # lưu dần để an toàn khi chạy dài
        save_status_store(STATUS_STORE_PATH, status_store)

    print(f"✅ DONE sheet {sheet_name} — output: {output_path} — errors: {error_path}")
    # save lần cuối (phòng hờ)
    save_status_store(STATUS_STORE_PATH, status_store)

# =========================
# RUN
# =========================
if __name__ == "__main__":
    # Khởi tạo Chrome profile thật (attach remote-debugging) — NHƯ CORE CỦA BẠN
    d = start_driver(CHROME_PATH, USER_DATA_DIR, PROFILE_NAME, port=REMOTE_PORT)
    # d = start_driver_with_proxy(PROXY_URL, headless=True)
    d.set_script_timeout(40)
    try:
        d.execute_cdp_cmd("Network.enable", {})
        d.execute_cdp_cmd("Network.setCacheDisabled", {"cacheDisabled": True})
    except Exception:
        pass

    # cài hook sớm (giống core)
    install_early_hook(d)

    # Nếu muốn, có thể bootstrap login ở ngoài (tuỳ bạn)
    # bootstrap_auth(d)

    crawl_from_excel_stream(
        INPUT_EXCEL,
        SHEET_NAME,
        OUTPUT_EXCEL,
        ERROR_EXCEL,
        driver=d,
        max_retries=0,
    )
    # d.quit()

import pandas as pd
from openpyxl import load_workbook
import traceback
import time
EXCEL_PATH = r"E:\NCS\fb-selenium\thoibao-de-last-split-sheet1.xlsx"
SHEET_NAME = "Sheet_1"   # <-- bạn có thể thay sheet khác
from get_comment_fb_automation import hook_graphql, install_early_hook, start_driver
from main import crawl_comments
from get_comment_fb_utils import open_reel_comments_if_present, set_sort_to_all_comments_unified
from configs import CHROME_PATH, USER_DATA_DIR, PROFILE_NAME, REMOTE_PORT
def read_links(sheet_name=SHEET_NAME):
    wb = load_workbook(EXCEL_PATH)
    if sheet_name not in wb.sheetnames:
        print(f"⚠️ Sheet {sheet_name} không tồn tại!")
        return pd.DataFrame()
    df = pd.read_excel(EXCEL_PATH, sheet_name=sheet_name)
    if 'link' not in df.columns:
        raise ValueError("Thiếu cột 'link' trong Excel.")
    if 'status' not in df.columns:
        df['status'] = ''
    return df, wb

def update_status(wb, sheet_name, row_index, status):
    ws = wb[sheet_name]
    # Excel dòng 1 là header, nên cộng thêm 2 để đến đúng dòng data
    ws.cell(row=row_index + 2, column=ws.max_column, value=status)
    wb.save(EXCEL_PATH)

def crawl_from_excel():
    df, wb = read_links(SHEET_NAME)
    print(f"📋 Tổng {len(df)} link, đang xử lý...")

    for idx, row in df.iterrows():
        url = str(row['link']).strip()
        status = str(row.get('status', '')).strip().lower()

        if not url or status == 'done':
            print(f"⏭️ Bỏ qua dòng {idx+2} (đã done hoặc trống).")
            continue

        print(f"\n🚀 Đang crawl link {idx+2}: {url}")

        try:
            # === chạy trình duyệt & crawl từng link ===
            d = start_driver(CHROME_PATH, USER_DATA_DIR, PROFILE_NAME, port=REMOTE_PORT)
            install_early_hook(d)
            d.get(url)
            time.sleep(2)
            hook_graphql(d)
            time.sleep(1)
            if "reel" in url:
                open_reel_comments_if_present(d)
            set_sort_to_all_comments_unified(d)
            time.sleep(1)

            crawl_comments(
                d,
                out_json=f"comments_{idx+1}.ndjson",
                checkpoint_path=f"checkpoint_{idx+1}.json",
                max_pages=None
            )

            update_status(wb, SHEET_NAME, idx, "done")
            print(f"✅ DONE link {idx+2}")

        except Exception as e:
            print(f"❌ FAIL link {idx+2}: {e}")
            traceback.print_exc()
            update_status(wb, SHEET_NAME, idx, "fail")

        finally:
            try:
                d.quit()
            except:
                pass

if __name__ == "__main__":
    crawl_from_excel()
